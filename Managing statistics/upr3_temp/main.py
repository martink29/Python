import os
import math
import time
import requests
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel  # noqa – само за справка
from openpyxl.chart import Series              # правилният импорт

import matplotlib
matplotlib.use("TkAgg")          # PyCharm Plots – смени на "TkAgg" ако Qt5 липсва
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker


# =============================================================================
#  НАСТРОЙКИ
# =============================================================================

LATITUDE   = 42.15
LONGITUDE  = 24.75
TIMEZONE   = "Europe/Sofia"

START_YEAR = 1976
END_YEAR   = 2026
MONTH      = 2

OUTPUT_FILE = "Февруари_Пловдив_Температури.xlsx"

TEMP_MIN  = -15.0
TEMP_MAX  =  24.5
TEMP_STEP =   0.5


# =============================================================================
#  СТЪПКА 1: ИЗТЕГЛЯНЕ НА ДАННИТЕ
# =============================================================================

def get_february_days(year: int) -> int:
    """Връща броя дни в февруари за дадена година."""
    return 29 if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0) else 28


def fetch_year_february(year: int, retries: int = 3) -> list:
    """
    Изтегля дневните средни температури за февруари на дадена година
    от Open-Meteo Historical Weather API.
    """
    days_in_feb = get_february_days(year)
    start_date  = f"{year}-02-01"
    end_date    = f"{year}-02-{days_in_feb:02d}"

    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude":   LATITUDE,
        "longitude":  LONGITUDE,
        "start_date": start_date,
        "end_date":   end_date,
        "daily":      "temperature_2m_mean",
        "timezone":   TIMEZONE,
    }

    for attempt in range(1, retries + 1):
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            temps = data["daily"]["temperature_2m_mean"]
            return [round(float(t), 1) if t is not None else None for t in temps]
        except requests.exceptions.RequestException as e:
            if attempt < retries:
                print(f"Опит {attempt}/{retries} неуспешен. Изчакване 3 сек...")
                time.sleep(3)
            else:
                print(f"Грешка при изтегляне за {year}: {e}")
                return [None] * days_in_feb

    return [None] * days_in_feb   # ← задоволява „missing return" предупреждението


def download_all_data(start_year: int, end_year: int):
    print(f"\n🌐 Изтегляме данни от Open-Meteo API")
    print(f"   Местоположение: Пловдив (lat={LATITUDE}, lon={LONGITUDE})")
    print(f"   Период: {start_year}–{end_year}, месец: Февруари\n")

    years = list(range(start_year, end_year + 1))
    data  = {yr: [None] * 29 for yr in years}

    for yr in years:
        print(f"  {yr}...", end=" ", flush=True)
        temps = fetch_year_february(yr)
        for i, t in enumerate(temps):
            data[yr][i] = t
        print(f"✓  ({len([t for t in temps if t is not None])} дни)")
        time.sleep(0.3)

    df = pd.DataFrame(data)
    df.insert(0, "Ден/Година", range(1, 30))

    all_vals = np.array([v for v in df[years].values.flatten()
                         if v is not None and not math.isnan(float(v))], dtype=float)
    print(f"\n✅ Данните са изтеглени успешно!")
    print(f"   Средна температура: {np.mean(all_vals):.1f}°C")
    print(f"   Диапазон: {np.min(all_vals):.1f}°C до {np.max(all_vals):.1f}°C")

    return df, years


# =============================================================================
#  СТЪПКА 2: СТАТИСТИКИ
# =============================================================================

def calculate_statistics(df, years: list) -> dict:
    all_values = np.array([float(v) for v in df[years].values.flatten()
                           if v is not None and not math.isnan(float(v))])

    mean_val   = float(np.mean(all_values))
    median_val = float(np.median(all_values))
    min_val    = float(np.min(all_values))
    max_val    = float(np.max(all_values))
    std_val    = float(np.std(all_values))
    mode_res   = scipy_stats.mode(np.round(all_values, 1), keepdims=True)
    mode_val   = float(mode_res.mode[0])
    beta_val   = std_val * 0.7797
    mu_val     = mean_val - 0.5772 * beta_val

    stats = {
        "Средна стойност":       round(mean_val,   4),
        "Медиана":               round(median_val, 4),
        "Минимална стойност":    round(min_val,    4),
        "Максимална стойност":   round(max_val,    4),
        "Стандартно отклонение": round(std_val,    4),
        "Мода":                  round(mode_val,   4),
        "Мащаб(Бета)":           round(beta_val,   4),
        "Локация(Ми)":           round(mu_val,     4),
    }

    print("\n Статистики:")
    for k, v in stats.items():
        print(f"   {k:<28} = {v}")

    return stats


# =============================================================================
#  СТЪПКА 3: РАЗПРЕДЕЛЕНИЯ
# =============================================================================

def calculate_distributions(stats: dict):
    mean  = stats["Средна стойност"]
    std   = stats["Стандартно отклонение"]
    a     = stats["Минимална стойност"]
    b     = stats["Максимална стойност"]
    c     = stats["Мода"]
    beta  = stats["Мащаб(Бета)"]
    mu    = stats["Локация(Ми)"]

    temps = []
    t = TEMP_MIN
    while t <= TEMP_MAX + 1e-9:
        temps.append(round(t, 1))
        t += TEMP_STEP

    normal_pdf     = []
    triangular_pdf = []
    gumbel_pdf     = []

    for t in temps:
        normal_pdf.append(round(scipy_stats.norm.pdf(t, loc=mean, scale=std), 8))

        if t < a or t > b or a == b:
            tri = 0.0
        elif t < c:
            denom = (b - a) * (c - a)
            tri = 2 * (t - a) / denom if denom != 0 else 0.0
        else:
            denom = (b - a) * (b - c)
            tri = 2 * (b - t) / denom if denom != 0 else 0.0
        triangular_pdf.append(round(tri, 8))

        if beta != 0:
            z     = (t - mu) / beta
            g_pdf = (1 / beta) * math.exp(-(z + math.exp(-z)))
        else:
            g_pdf = 0.0
        gumbel_pdf.append(round(g_pdf, 8))

    print(f"\nРазпределения изчислени за {len(temps)} температурни точки")
    return temps, normal_pdf, triangular_pdf, gumbel_pdf


# =============================================================================
#  СТЪПКА 4: MATPLOTLIB ГРАФИКИ (появяват се в PyCharm Plots)
# =============================================================================

def plot_distributions(temps: list, normal_pdf: list,
                       triangular_pdf: list, gumbel_pdf: list) -> None:
    """
    Рисува 4 графики в PyCharm Plots (matplotlib):
      1. Нормално разпределение
      2. Триъгълно разпределение
      3. Разпределение на Гъмбел
      4. Трите заедно
    """
    print("\nСъздаваме matplotlib графики...")

    x = temps

    # Цветове (съвпадат с Excel)
    COLOR_NORM  = "#4285F4"   # синьо
    COLOR_TRI   = "#EA4335"   # червено
    COLOR_GUMB  = "#FBBC04"   # жълто

    fig, axes = plt.subplots(2, 2, figsize=(14, 9))
    fig.suptitle("Температури Февруари – Пловдив (1976–2026)\nРазпределения",
                 fontsize=14, fontweight="bold", y=1.01)

    plots_cfg = [
        (axes[0, 0], "Нормално разпределение",   normal_pdf,     COLOR_NORM),
        (axes[0, 1], "Триъгълно разпределение",   triangular_pdf, COLOR_TRI),
        (axes[1, 0], "Разпределение на Гъмбел",   gumbel_pdf,     COLOR_GUMB),
        (axes[1, 1], None,                         None,           None),   # комбинирана
    ]

    for ax, title, y_data, color in plots_cfg:
        if title is not None:
            ax.plot(x, y_data, "o", color=color, markersize=4, label=title)
            ax.fill_between(x, y_data, alpha=0.15, color=color)
            ax.set_title(title, fontsize=11, fontweight="bold")
        else:
            # Комбинирана графика
            ax.plot(x, normal_pdf,     "o", color=COLOR_NORM, markersize=3,
                    label="Нормално")
            ax.plot(x, triangular_pdf, "o", color=COLOR_TRI,  markersize=3,
                    label="Триъгълно")
            ax.plot(x, gumbel_pdf,     "o", color=COLOR_GUMB, markersize=3,
                    label="Гъмбел")
            ax.fill_between(x, normal_pdf,     alpha=0.10, color=COLOR_NORM)
            ax.fill_between(x, triangular_pdf, alpha=0.10, color=COLOR_TRI)
            ax.fill_between(x, gumbel_pdf,     alpha=0.10, color=COLOR_GUMB)
            ax.set_title("Разпределения – обща диаграма",
                         fontsize=11, fontweight="bold")
            ax.legend(fontsize=9)

        ax.set_xlabel("Температура (°C)", fontsize=9)
        ax.set_ylabel("Вероятност (плътност)", fontsize=9)
        ax.xaxis.set_major_locator(mticker.MultipleLocator(5))
        ax.grid(True, linestyle="--", alpha=0.4)
        ax.tick_params(labelsize=8)

    plt.tight_layout()
    plt.show()
    print("  Графиките са показани в PyCharm Plots.")


# =============================================================================
#  СТЪПКА 5: EXCEL С ГРАФИКИ
# =============================================================================

def write_excel(df, years: list, stats: dict, temps: list,
                normal_pdf: list, triangular_pdf: list,
                gumbel_pdf: list, output_file: str) -> None:

    print(f"\nСъздаваме Excel файл: {output_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    yellow_fill  = PatternFill("solid", fgColor="FFFF00")
    blue_fill    = PatternFill("solid", fgColor="DDEEFF")
    green_fill   = PatternFill("solid", fgColor="E2EFDA")
    header_font  = Font(name="Arial", bold=True, size=11)
    normal_font  = Font(name="Arial", size=11)
    stats_font   = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")

    def style_header(cell, fill=yellow_fill):
        cell.font      = header_font
        cell.fill      = fill
        cell.alignment = center_align

    # --- Заглавен ред ---
    header_cell = ws.cell(row=1, column=1, value="Ден/Година")
    style_header(header_cell)
    ws.column_dimensions["A"].width = 14

    for j, yr in enumerate(years):
        col  = j + 2
        yr_cell = ws.cell(row=1, column=col, value=yr)
        style_header(yr_cell)
        ws.column_dimensions[get_column_letter(col)].width = 7

    # --- Данни ---
    for d_idx in range(29):
        row       = d_idx + 2
        day_cell  = ws.cell(row=row, column=1, value=float(d_idx + 1))
        style_header(day_cell)
        for j, yr in enumerate(years):
            val = df.at[d_idx, yr]
            if val is not None and not (isinstance(val, float) and math.isnan(val)):
                data_cell           = ws.cell(row=row, column=j + 2, value=float(val))
                data_cell.font      = normal_font
                data_cell.alignment = center_align

    # --- Статистики ---
    last_data_col  = len(years) + 1
    stat_label_col = last_data_col + 2
    stat_value_col = last_data_col + 3

    for i, (label, value) in enumerate(stats.items()):
        row    = 6 + i
        lc     = ws.cell(row=row, column=stat_label_col, value=label)
        vc     = ws.cell(row=row, column=stat_value_col, value=value)
        lc.font = stats_font;  lc.fill = green_fill
        vc.font = stats_font;  vc.fill = green_fill
        vc.number_format = "0.0000"

    ws.column_dimensions[get_column_letter(stat_label_col)].width = 24
    ws.column_dimensions[get_column_letter(stat_value_col)].width = 14

    # --- Таблица с разпределения ---
    dist_start_col = last_data_col + 7
    dist_start_row = 5

    dist_headers = [
        "Температура",
        "Нормално разпределение",
        "Триъгълно разпределение",
        "Разпределение на Гъмбел",
    ]
    for k, hdr in enumerate(dist_headers):
        col    = dist_start_col + k
        hdr_cell = ws.cell(row=dist_start_row, column=col, value=hdr)
        style_header(hdr_cell, fill=blue_fill)
        ws.column_dimensions[get_column_letter(col)].width = 26 if k > 0 else 13

    dist_data_start = dist_start_row + 1
    dist_data_end   = dist_start_row + len(temps)

    for i, t in enumerate(temps):
        row = dist_data_start + i
        ws.cell(row=row, column=dist_start_col,     value=t).font                = normal_font
        ws.cell(row=row, column=dist_start_col + 1, value=normal_pdf[i]).font    = normal_font
        ws.cell(row=row, column=dist_start_col + 2, value=triangular_pdf[i]).font = normal_font
        ws.cell(row=row, column=dist_start_col + 3, value=gumbel_pdf[i]).font    = normal_font

    temp_col = dist_start_col
    norm_col = dist_start_col + 1
    tri_col  = dist_start_col + 2
    gumb_col = dist_start_col + 3

    # --- Графики в Excel ---
    def make_series(x_col: int, y_col: int,
                    d_start: int, d_end: int,
                    ser_title: str, color: str) -> Series:
        xvals = Reference(ws, min_col=x_col, min_row=d_start, max_row=d_end)
        yvals = Reference(ws, min_col=y_col, min_row=d_start, max_row=d_end)
        ser = Series(yvals, xvals, title=ser_title)
        ser.marker.symbol                              = "circle"
        ser.marker.size                                = 5
        ser.marker.graphicalProperties.solidFill       = color
        ser.marker.graphicalProperties.line.solidFill  = color
        ser.graphicalProperties.line.noFill            = True
        return ser

    charts_cfg = [
        {
            "title":   "Нормално разпределение на температурите",
            "x_label": "Температура (°C)", "y_label": "Вероятност (Плътност)",
            "series":  [("Нормално разпределение",  norm_col, "4285F4")],
            "anchor":  f"{get_column_letter(dist_start_col)}36",
            "w": 18, "h": 12,
        },
        {
            "title":   "Триъгълно разпределение на температурите",
            "x_label": "Температура (°C)", "y_label": "Вероятност (Плътност)",
            "series":  [("Триъгълно разпределение", tri_col,  "EA4335")],
            "anchor":  f"{get_column_letter(dist_start_col + 4)}36",
            "w": 18, "h": 12,
        },
        {
            "title":   "Разпределение на Гъмбел",
            "x_label": "Температура (°C)", "y_label": "Вероятност (Плътност)",
            "series":  [("Разпределение на Гъмбел",  gumb_col, "FBBC04")],
            "anchor":  f"{get_column_letter(dist_start_col)}62",
            "w": 18, "h": 12,
        },
        {
            "title":   "Разпределения обща диаграма",
            "x_label": "Температура (°C)", "y_label": "Вероятност (Плътност)",
            "series":  [
                ("Нормално разпределение",  norm_col, "4285F4"),
                ("Триъгълно разпределение", tri_col,  "EA4335"),
                ("Разпределение на Гъмбел", gumb_col, "FBBC04"),
            ],
            "anchor":  f"{get_column_letter(dist_start_col + 4)}62",
            "w": 22, "h": 14,
        },
    ]

    for cfg in charts_cfg:
        chart                = ScatterChart()
        chart.title          = cfg["title"]
        chart.style          = 10
        chart.x_axis.title   = cfg["x_label"]
        chart.y_axis.title   = cfg["y_label"]
        chart.legend.position = "b"
        chart.width          = cfg["w"]
        chart.height         = cfg["h"]
        for ser_title, y_col_idx, color in cfg["series"]:
            chart.series.append(
                make_series(temp_col, y_col_idx,
                            dist_data_start, dist_data_end,
                            ser_title, color)
            )
        ws.add_chart(chart, cfg["anchor"])
        print(f"   Графика: '{cfg['title']}'")

    wb.save(output_file)
    print(f"\nЗаписан: {output_file}")
    print(f"   Пълен път: {os.path.abspath(output_file)}")


# =============================================================================
#  ГЛАВНА ФУНКЦИЯ
# =============================================================================

def main() -> None:
    print("=" * 65)
    print("  ТЕМПЕРАТУРИ ФЕВРУАРИ – ПЛОВДИВ  |  Данни от интернет")
    print("=" * 65)

    # 1. Изтегляне
    df, years = download_all_data(START_YEAR, END_YEAR)

    # 2. Статистики
    stats = calculate_statistics(df, years)

    # 3. Разпределения
    temps, normal_pdf, tri_pdf, gumbel_pdf = calculate_distributions(stats)

    # 4. Matplotlib графики – появяват се ВЕДНАГА в PyCharm Plots
    plot_distributions(temps, normal_pdf, tri_pdf, gumbel_pdf)

    # 5. Excel с вградени графики
    write_excel(df, years, stats, temps, normal_pdf, tri_pdf,
                gumbel_pdf, OUTPUT_FILE)

    print("\n Готово! Отвори файла в Excel.")
    print("=" * 65)


if __name__ == "__main__":
    main()